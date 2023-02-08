from datetime import datetime 
from openpyxl import load_workbook
Rut=r"C:\Users\DELL\Desktop\BaseCrud.xlsx"

def leer(ruta:str, extraer:str):
    Archivo_Excel = load_workbook(ruta)
    Hojas_datos = Archivo_Excel['DatosCrud']
    Hojas_datos= Hojas_datos ['A':'F'+str(Hojas_datos.max_row)]
    info={}

    for i in Hojas_datos:

        if isinstance(i[0].value, int):
            info.setdefault(i[0].value,{'tarea':i[1].value, 'descripcion':i[2].value,
                                         'estado':i[3].value, 'fecha de inicio':i[4].value,
                                         'fecha de finalizacion':i[5].value,})
    if not (extraer =='todo'):
        info=filtrar(info, extraer)

    for i in info:
        print('******Tarea******')
        print('Id:'+ str(i)+'\n'+'Titulo¨: '+str(info[i]['tarea'])+'\n'+'Descripcion: '
        +str(info[i]['descripcion'])+'\n'+'Estado: '+str(info[i]['estado'])
        +'\n'+'Fecha de Creacion: '+str(info[i]['fecha de inicio'])
        + +'\n'+'Fecha de finalizacion: '+str(info[i]['fecha de finalizacion']))
        print()
    
    return

def filtrar(info:dict,filtro:str):
    aux={}
    for i in info:
        if info[i]['estado']==filtro:
            aux.setdefault(i,info[i])
    return aux


def actualizar(ruta: str, identicador: int,datos_actualizados:dict):
    Archivo_Excel = load_workbook(ruta)
    Hoja_datos = Archivo_Excel['DatosCrud']
    Hoja_datos=Hoja_datos['A2':'F' +str(Hoja_datos.max_row)]
    hoja=Archivo_Excel.active

    Titulo=2
    Descripcion=3
    Estado=4
    Fecha_Inicio=5
    Fecha_Fin=6
    encontro=False
    for i in Hoja_datos:
        if i[0].value==identicador:
            fila=i[0].row
            encontro=True
            for d in datos_actualizados:
                if d == 'titulo' and not (datos_actualizados[d]==''):
                    hoja.cell(row=fila, column=Titulo).value=datos_actualizados[d]
                elif  d == 'descripcion' and not (datos_actualizados[d]==''):
                     hoja.cell(row=fila, column=Descripcion).value=datos_actualizados[d]
                elif  d == 'estado' and not (datos_actualizados[d]==''):
                     hoja.cell(row=fila, column=Estado).value=datos_actualizados[d]
                elif  d == 'fecha inicio' and not (datos_actualizados[d]==''):
                     hoja.cell(row=fila, column=Fecha_Inicio).value=datos_actualizados[d]
                elif  d == 'fecha finalizacion' and not (datos_actualizados[d]==''):
                     hoja.cell(row=fila, column=Fecha_Fin).value=datos_actualizados[d]
    Archivo_Excel.save(ruta)
    if encontro==False:
        print('Error: No existe una tarea con es Id')
        print()
    return


def agregar(ruta: str, datos:dict):
    Archivo_Excel = load_workbook(ruta)
    Hoja_datos = Archivo_Excel['DatosCrud']
    Hoja_datos=Hoja_datos['A2':'F' +str(Hoja_datos.max_row+1)]
    hoja=Archivo_Excel.active

    Titulo=2
    Descripcion=3
    Estado=4
    Fecha_Inicio=5
    Fecha_Fin=6
    encontro=False
    for i in Hoja_datos:

        if not (isinstance(i[0].value, int)):
            identicador=1[0].row
            hoja.cell(row=identicador, column=1).value=identicador-1
            hoja.cell(row=identicador, column=Titulo).value=datos['titulo']
            hoja.cell(row=identicador, column=Descripcion).value=datos['desacripcion']
            hoja.cell(row=identicador, column=Estado).value=datos['estado']
            hoja.cell(row=identicador, column=Fecha_Inicio).value=datos['fecha inicio']
            hoja.cell(row=identicador, column=Fecha_Fin).value=datos['fecha finalizacion']
            break
    Archivo_Excel.save(ruta)
    return

def borrar(ruta, identificador):
    Archivo_Excel = load_workbook(ruta)
    Hoja_datos = Archivo_Excel['DatosCrud']
    Hoja_datos=Hoja_datos['A2':'F' +str(Hoja_datos.max_row)]
    hoja=Archivo_Excel.active

    Titulo=2
    Descripcion=3
    Estado=4
    Fecha_Inicio=5
    Fecha_Fin=6
    encontro=False
    for i in Hoja_datos:
        if i[0].value==identificador:
            fila=i[0].row
            encontro=True

            hoja.cell(row=fila, column=1).value=""
            hoja.cell(row=fila, column=Titulo).value=""
            hoja.cell(row=fila, column=Descripcion).value=""
            hoja.cell(row=fila, column=Estado).value=""
            hoja.cell(row=fila, column=Fecha_Inicio).value=""
            hoja.cell(row=fila, column=Fecha_Fin).value=""
    Archivo_Excel.save(ruta)
    if encontro==False:
        print('Error: No existe una tarea con es Id')
        print()
    return